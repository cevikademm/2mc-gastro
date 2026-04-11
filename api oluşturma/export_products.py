"""
Diamond EU — Tüm Ürün Verilerini Excel'e Aktar
"""

import requests
import json
import sys
import re
import pandas as pd
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import os
from dotenv import load_dotenv
load_dotenv()

BASE_URL = "https://api.diamond-eu.com"
EMAIL    = os.getenv("DIAMOND_EMAIL", "")
PASSWORD = os.getenv("DIAMOND_PASSWORD", "")
LANGUAGE = os.getenv("DIAMOND_LANGUAGE", "tr")


def login():
    print("🔐 Giriş yapılıyor...")
    r = requests.post(f"{BASE_URL}/auth/login",
                      json={"email": EMAIL, "password": PASSWORD})
    r.raise_for_status()
    token = r.json()["access_token"]
    print("✅ Giriş başarılı!\n")
    return token


def fetch_all_products(token):
    print("📦 Ürünler çekiliyor...")
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept-Language": LANGUAGE
    }
    all_products = []
    url = f"{BASE_URL}/products-export?filter[is_old][value]=0"

    while url:
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
        products = data.get("data", [])
        all_products.extend(products)
        print(f"   → {len(all_products)} ürün alındı...")
        url = data.get("links", {}).get("next")

    print(f"✅ Toplam {len(all_products)} ürün çekildi.\n")
    return all_products


def fetch_menu_trees(token):
    print("🌳 Kategori isimleri çekiliyor...")
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept-Language": LANGUAGE
    }
    r = requests.get(f"{BASE_URL}/menu-trees", headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    # Kategori ID → isim eşleştirmesi
    mapping = {}
    for item in data.get("included", []):
        item_id = item.get("id")
        name = item.get("attributes", {}).get("name", "")
        if item_id and name:
            mapping[str(item_id)] = name

    print(f"✅ {len(mapping)} kategori ismi alındı.\n")
    return mapping


def parse_products(products, category_map):
    print("📊 Veriler düzenleniyor...")
    rows = []

    for p in products:
        attr = p.get("attributes", {})
        price = attr.get("price", {})
        media = attr.get("media", {})
        images = media.get("images", [])

        # Görseller
        img_big   = images[0].get("Big", "") if images else ""
        img_thumb = images[0].get("thumb", "") if images else ""
        img_gallery = images[0].get("Thumb-gallery", "") if images else ""
        img_full  = images[0].get("full", "") if images else ""

        # Kategori isimleri
        cat_id      = str(attr.get("product_category_id", ""))
        range_id    = str(attr.get("product_range_id", ""))
        subrange_id = str(attr.get("product_subrange_id", "") or "")
        family_id   = str(attr.get("product_family_id", ""))
        subfamily_id = str(attr.get("product_subfamily_id", "") or "")
        line_id     = str(attr.get("product_line_id", "") or "")

        row = {
            # KİMLİK
            "Ürün ID":           p.get("id", ""),
            "Ürün Adı":          attr.get("name", ""),
            "Teknik Açıklama":   attr.get("description_tech_spec", ""),
            "Ek Bilgi":          attr.get("popup_info", ""),

            # FİYAT
            "Para Birimi":       price.get("currency", "EUR"),
            "Katalog Fiyatı":    price.get("catalog", ""),
            "Görüntülenen Fiyat": price.get("display", ""),
            "Promosyon Fiyatı":  price.get("promo", ""),
            "Katalog Sayfa No":  attr.get("page_catalog_number", ""),
            "Promo Sayfa No":    attr.get("page_promo_number", ""),

            # STOK
            "Stok Miktarı":      attr.get("stock", ""),
            "Yeniden Stoklama":  attr.get("restock_info", ""),
            "Tedarikçi Teslimat (gün)": attr.get("supplier_delivery_delay", ""),
            "Ort. Stoklama Süresi (gün)": attr.get("days_to_restock_avg", ""),

            # BOYUT / AĞIRLIK
            "Uzunluk (mm)":      attr.get("length_mm", ""),
            "Genişlik (mm)":     attr.get("width_mm", ""),
            "Yükseklik (mm)":    attr.get("height_mm", ""),
            "Hacim (m³)":        attr.get("volume_m3", ""),
            "Ağırlık":           attr.get("weight", ""),
            "Ağırlık Birimi":    attr.get("weight_unit", ""),

            # TEKNİK
            "Elektrik Gücü (kW)":  attr.get("electric_power_kw", ""),
            "Elektrik Bağlantı":   attr.get("electric_connection", ""),
            "Elektrik Bağlantı 2": attr.get("electric_connection_2", ""),
            "Buhar":               attr.get("vapor", ""),
            "Kcal Gücü":          attr.get("kcal_power", ""),
            "Beygir Gücü":        attr.get("horse_power", ""),

            # KATEGORİ
            "Kategori ID":       cat_id,
            "Kategori Adı":      category_map.get(cat_id, ""),
            "Ürün Grubu ID":     range_id,
            "Ürün Grubu Adı":    category_map.get(range_id, ""),
            "Alt Grup ID":       subrange_id,
            "Alt Grup Adı":      category_map.get(subrange_id, ""),
            "Ürün Ailesi ID":    family_id,
            "Ürün Ailesi Adı":   attr.get("product_family_name", "") or category_map.get(family_id, ""),
            "Alt Aile ID":       subfamily_id,
            "Alt Aile Adı":      category_map.get(subfamily_id, ""),
            "Ürün Hattı ID":     line_id,
            "Ürün Hattı Adı":    category_map.get(line_id, ""),

            # MEDYA
            "Görsel (Büyük)":    img_big,
            "Görsel (Küçük)":    img_thumb,
            "Görsel (Galeri)":   img_gallery,
            "Görsel (Full)":     img_full,

            # DURUM
            "Yeni Ürün":        "Evet" if attr.get("is_new") else "Hayır",
            "Eski Ürün":        "Evet" if attr.get("is_old") else "Hayır",
            "Kampanyalı":       "Evet" if attr.get("is_good_deal") else "Hayır",
            "Ürün Tipi":        attr.get("product_type", ""),
            "Aksesuar Var":     "Evet" if attr.get("has_accessories") else "Hayır",
            "Yedek Ürün ID":    attr.get("replacement_product_id", ""),
        }
        rows.append(row)

    # Excel'de geçersiz karakterleri temizle
    illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
    for row in rows:
        for key, val in row.items():
            if isinstance(val, str):
                row[key] = illegal_chars.sub("", val)

    print(f"✅ {len(rows)} ürün düzenlendi.\n")
    return rows


def save_to_excel(rows, filename):
    print(f"💾 Excel'e yazılıyor: {filename}")
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tüm Ürünler", index=False, freeze_panes=(1, 0))

        wb = writer.book
        ws = wb["Tüm Ürünler"]

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            header_len = len(str(col[0].value or ""))
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len, header_len) + 2, 50)

        # Başlık satırına filtre ekle
        ws.auto_filter.ref = ws.dimensions

        # Başlık stilini ayarla
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    print(f"✅ Excel kaydedildi: {filename}")
    print(f"   📊 {len(rows)} satır, {len(rows[0]) if rows else 0} sütun\n")


if __name__ == "__main__":
    # 1. Giriş
    token = login()

    # 2. Kategori isimlerini çek
    category_map = fetch_menu_trees(token)

    # 3. Ürünleri çek
    products = fetch_all_products(token)

    # 4. Verileri düzenle
    rows = parse_products(products, category_map)

    # 5. Excel'e kaydet
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Diamond_Urunler_{timestamp}.xlsx"
    save_to_excel(rows, filename)

    # 6. Ham JSON'u da kaydet (yedek)
    with open("products_raw.json", "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print("💾 Ham veri kaydedildi: products_raw.json")

    print("\n🎉 Tüm işlemler tamamlandı!")
